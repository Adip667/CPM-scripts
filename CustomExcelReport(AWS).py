import csv
import os
import urllib3 as urllib3
from openpyxl import Workbook, load_workbook
from time import strftime
import datetime
from utils import call_rest_endpoint, get_token,print_all

##########
### PLEASE UPDATE - required input for demo ###
##########
host = '1.1.1.1'  # used for all steps in the demo
api_key = '1f0240e8ce04bd378a7705d4b91f1cdcd1ef50324'  # used in step 1 to generate access token
user = 1  # used in steps 3 -> 12, the user ID to create the report for.
report_name_prefix = 'MyCustomReport_'  # used in step 2 to create the file name
num_days = 7  # used in step 8 to calculate the from_date filter for the audit/snapshot/backup report generation.
##########
##########

def create_xlsx(**kwargs):
    '''
    Create the intial xlsx file with the relevant tabs.
    :param kwargs['name']: will contain the file name to be created
    '''
    wb = Workbook()
    wb.create_sheet(title='General Information')
    wb.create_sheet(title='Audit')
    wb.create_sheet(title='Backup Records')
    wb.create_sheet(title='Snapshots')
    wb.create_sheet(title='Recoveries')
    del wb['Sheet']  # Remove the default sheet as we only want the tabs we defined above
    wb.save(kwargs['name'])  # save the created file


def _add_row_xlsx(**kwargs):
    '''
    Add row to the xlsx file created with create_xlsx(), this function used only by add_section_xlsx()
    :param kwargs['report_name']: the name of the xlsx file
    :param kwargs['sheet_name']: the name of the tab the row need to be added too.
    :param kwargs['row']: the row to be added
    '''
    wb = load_workbook(kwargs['report_name'])  # open the xlsx file
    ws = wb[kwargs['sheet_name']]  # set relevant tab
    ws.append(kwargs['row'])  # add row
    wb.save(kwargs['report_name'])  # save


def add_section_xlsx(responses, **kwargs):
    '''
    Add data as a table to the relevant tab, for example the list of policies
    :param responses: contain all the data to be added to report.
    :param kwargs['name']: contain the table name to be added, for example policies or users.
    :param kwargs['sheet_name']: the excel tab the data need to be added too.
    :param kwargs['table_headers']: the table header for the data, for examle Id, name, etc...
    :param kwargs['keys']: list of the relevant keys for the above responses param.
    '''
    row = (kwargs['name'] + ' ({})'.format(len(responses)),)  # create the table title with the number of items, for example: Polices(39)
    _add_row_xlsx(report_name=report_name, sheet_name=kwargs['sheet_name'], row=row)  # add above row to report.
    for n in range(len(responses)):  # if there are items, create table with info
        row = []  # will contain the row to be added to the excel
        if n == 0:  # first iteration, we need to create the table headers, for example id,name etc..
            _add_row_xlsx(report_name=report_name, sheet_name=kwargs['sheet_name'], row=kwargs['table_headers'])  # add kwargs['table_headers'] as a row to excel
        for key in kwargs['keys']:  # we will build the row dynamically based on responses & kwargs['keys']
            if type(key) == str: # if the key contain the data, for example: {'id': 2. etc..}
                if type(responses[n].get(key)) == int:
                    row.append(responses[n].get(key))
                else:
                    row.append(str(responses[n].get(key)))
            else:  # if key contain another dictionary in it that has the data, like in step 3 - for example: {'profile': {'allow_recovery': True}, etc..}
                key_name = list(key.keys())[0]
                row.append(responses[n][key_name][key[key_name]])
        _add_row_xlsx(report_name=report_name, sheet_name=kwargs['sheet_name'], row=row)  # add the row data we dynamically constructed to the excel
    _add_row_xlsx(report_name=report_name, sheet_name=kwargs['sheet_name'], row=(('',)))  # add empty row to separate in case we run this function again.

def add_csv_to_xlsx_report(**kwargs):
    '''
    Add csv report content to the xlsx report
    :param kwargs['report_name']: the name of the xlsx file
    :param kwargs['sheet_name']: the excel tab the csv data need to be added too.
    :param kwargs['csv_name']: name of the csv file
    '''
    wb = load_workbook(kwargs['report_name'])  # open file
    ws = wb[kwargs['sheet_name']]  # the tab to which we want to add the csv report
    with open(kwargs['csv_name']) as f:  # read the file
        reader = csv.reader(f, delimiter=',')
        for row in reader:  # add each row in the csv
            ws.append(row)
    wb.save(kwargs['report_name'])  # save
    os.remove(kwargs['csv_name'])  # delete the csv file after we completed adding it to the excel.


if __name__ == '__main__':
    URL_API = 'https://{}/api/'.format(host)  # updating the base URL with the provide host
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

    print('\nRunning Example (created using: N2WS V4.0.0c, RESTful API Guide V2.0.0)')

    print('\nStep 1. Getting access token using API KEY and updating the headers')
    tokens = get_token(URL_API, api_key)  # calling function to generate access token from the provided API key
    headers_json = {'Accept': 'application/json;', 'Authorization': 'Bearer {}'.format(tokens[0])}  # creating json header with the access token, will be used later for get API calls such as list policies/user/accounts/etc
    headers_csv = {'Accept': 'text/csv;', 'Authorization': 'Bearer {}'.format(tokens[0])}  # creating csv header with the access token, will be used later for API calls to generate Audit/Backup/Snapshot report

    print('\nStep 2. Creating the empty xlsx report')
    report_name = strftime(report_name_prefix + '(%Y-%b-%d)_(%H-%M-%S).xlsx')  # connecting the report_name prefix with timestamp
    create_xlsx(name=report_name)  # creating the excel file, we will populate it later with all the information

    print('\nStep 3. Check user delegates and add to excel')
    url = URL_API + 'delegates/?original_user={}'.format(user)  # update url with relevant endpoint, API Used: 31.2. List CPM delegate users
    responses = call_rest_endpoint(url=url, headers=headers_json,type='get')  # calling the API using the util.py -> call_rest_endpoint(), this is a common function that will be used by all step in the demo
    print_all(responses)
    table_headers = (  # Were going to add the delegate user list as a table to the excel, this will be the table headers
        'ID', 'Name', 'Email', 'allow_recovery', 'allow_account_changes', 'allow_backup_changes','allow_settings_changes')
    keys = ('id', 'username', 'email', {'profile': 'allow_recovery'}, {'profile': 'allow_account_changes'}, {'profile': 'allow_backup_changes'},
            {'profile': 'allow_settings_changes'})  # keys in the responses dictionary we want to add to the excel
    add_section_xlsx(responses, name='User Delegates', sheet_name='General Information', table_headers=table_headers, keys=keys)  # add the table to the General Information tab in the excel

    print('\nStep 4. Check accounts for the user and add to excel')
    url = URL_API + 'aws/accounts/?user={}'.format(user)  # API Used: 5.1. List CPM accounts
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    print_all(responses)
    table_headers = ('ID', 'Name', 'DR Account', 'Auth Type', 'Capture vpcs')
    keys = ('id', 'name', 'is_dr_account', 'authentication', 'capture_vpcs')
    add_section_xlsx(responses, name='Accounts', sheet_name='General Information', table_headers=table_headers, keys=keys)

    print('\nStep 5. Check policies for the user and add to excel')
    url = URL_API + 'aws/policies/?user={}'.format(user)  # API Used: 10.2. List CPM policies
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    print_all(responses)
    table_headers = ('ID', 'Name', 'generations', 'enabled', 'schedules', 'dr_enabled', 'copy_to_s3_enabled', 'last_modified')
    keys = ('id', 'name', 'generations', 'enabled', 'schedules', 'dr_enabled', 'copy_to_s3_enabled', 'last_modified')
    add_section_xlsx(responses, name='Policies', sheet_name='General Information', table_headers=table_headers, keys=keys)

    print('\nStep 6. Check schedules and add to the excel')
    url = URL_API + 'schedules/?user={}'.format(user)  # API Used: 35.2. List CPM schedules
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    print_all(responses)
    table_headers = ('ID', 'Name', 'Time Unit', 'How often', 'last_modified')
    keys = ('id', 'name', 'every_unit', 'every_how_many', 'last_modified')
    add_section_xlsx(responses, name='schedules', sheet_name='General Information', table_headers=table_headers, keys=keys)

    print('\nStep 7. Check Resource Control groups and add tot the excel')
    url = URL_API + 'aws/resource_control/groups/?user={}'.format(user)  # API Used: 15.2. List CPM resource control groups
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    print_all(responses)
    table_headers = ('ID', 'Name', 'Mode', 'enabled', 'last_modified')
    keys = ('id', 'name', 'operation_mode', 'enabled', 'last_modified')
    add_section_xlsx(responses, name='Resource Control groups', sheet_name='General Information', table_headers=table_headers, keys=keys)

    print('\nStep 8. Check/Add Report schedule')
    url = URL_API + 'scheduled_reports/?user={}'.format(user)  # API Used: 34.2. List scheduled reports
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    print_all(responses)
    table_headers = ('ID', 'Name', 'enabled', 'Type', 'recipients', 'schedules')
    keys = ('id', 'name', 'enabled', 'report_type', 'recipients', 'schedules')
    add_section_xlsx(responses, name='Scheduled Reports', sheet_name='General Information', table_headers=table_headers, keys=keys)

    print('\nStep 9. get the backups report for the provided time(NUM_DAYS) and add to excel')
    from_date = (datetime.datetime.now() - datetime.timedelta(days=num_days)).strftime('%Y-%m-%dT%H:%M:%SZ')  # calculate the from_time based on the num_days provided
    print(' from_date={}'.format(from_date))  # print date just for demo purposes
    url = URL_API + 'reports/backups/?user={}&from_time={}'.format(user, from_date)  # API Used: 33.2. Backup report
    responses = call_rest_endpoint(url=url, headers=headers_csv, type='get', file=True)  # call the rest endpoint to get the report file
    add_csv_to_xlsx_report(report_name=report_name, sheet_name='Backup Records', csv_name=responses)  # add the whole csv file content to the Backup Records tab

    print('\nStep 10. get the snapshot report for the provided time(NUM_DAYS) and add to excel')
    url = URL_API + 'reports/snapshots/?user={}&from_time={}'.format(user, from_date)  # API Used: 33.6. Snapshots report
    responses = call_rest_endpoint(url=url, headers=headers_csv, type='get', file=True)
    add_csv_to_xlsx_report(report_name=report_name, sheet_name='Snapshots', csv_name=responses)  # add the whole csv file content to the snapshots tab

    print('\nStep 11. get the audit report for the provided time(NUM_DAYS) and add to excel')
    url = URL_API + 'reports/audit/?user={}&from_time={}'.format(user, from_date)  # API Used: 33.1. Audit report
    responses = call_rest_endpoint(url=url, headers=headers_csv, type='get', file=True)
    add_csv_to_xlsx_report(report_name=report_name, sheet_name='Audit', csv_name=responses)  # add the whole csv file content to the Audit tab

    print('\nStep 12. get recoveries list and add to excel')
    url = URL_API + 'aws/recoveries/?user={}&ordering=-recovery_time'.format(user)  # get the recoveries list for the user in descending order, API Used: 12.1. List CPM recovery records
    responses = call_rest_endpoint(url=url, headers=headers_json, type='get')
    for_xlsx_list = []  # A list that will hold only the records we want to add
    for item in responses:  # responses contain all the recoveries, so we will check for items newer then from_date and add to for_xlsx_list
        if item['recovery_time'] >= from_date:  # if item in the correct time, add to list
            print(' {}'.format(item))
            for_xlsx_list.append(item)
        else:  # recovery to old stop loop
            print(' too old(stopping loop):{}'.format(item))
            break
    table_headers = ('ID', 'recovery_time', 'backup_time', 'policy ID', 'Type', 'status', 'related recovery scenario id')
    keys = ('id', 'recovery_time', 'backup_time', 'policy', 'recovery_type', 'status','recovery_scenario_id')
    add_section_xlsx(for_xlsx_list, name='Recoveries', sheet_name='Recoveries', table_headers=table_headers, keys=keys)  # add recovery table to Recoveries tab
